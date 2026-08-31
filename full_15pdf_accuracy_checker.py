import os
import re
from datetime import datetime
from pdf2image import convert_from_path
import pytesseract
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font

# ====================== PATHS ======================
SCANNED_PDF_FOLDER = r"C:\Users\padma\OneDrive\Desktop\Anjuman\her office work august 28\Annual reports-20260804T103947Z-1-001\Annual reports"
EXCEL_PDF_PATH = r"C:\Users\padma\Downloads\divyajyoti-historical-dataset-v9.pdf"

TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH   = r"C:\poppler\Library\bin"

OUTPUT_FOLDER = r"C:\Users\padma\Documents\Full_15PDF_Accuracy_Report"
# ===================================================

pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
OCR_TEXT_FOLDER = os.path.join(OUTPUT_FOLDER, "OCR_Texts")
os.makedirs(OCR_TEXT_FOLDER, exist_ok=True)


def extract_numbers(text):
    pattern = r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?\b|\b\d+\.\d+\b|\b\d+\b'
    numbers = re.findall(pattern, text)
    cleaned = []
    for n in numbers:
        n_clean = n.replace(",", "")
        try:
            float(n_clean)
            cleaned.append(n_clean)
        except:
            pass
    return cleaned


def ocr_pdf(pdf_path, txt_path):
    print(f"   Converting pages to images...")
    images = convert_from_path(pdf_path, dpi=200, poppler_path=POPPLER_PATH)

    all_text = []
    for i, img in enumerate(images, 1):
        print(f"   OCR page {i}/{len(images)}", end="\r")
        text = pytesseract.image_to_string(img, lang="eng")
        all_text.append(f"\n\n----- PAGE {i} -----\n\n{text}")

    full_text = "".join(all_text)
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(full_text)
    print(f"\n   Saved OCR text")
    return full_text


def extract_text_from_clean_pdf(pdf_path):
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
    return "\n\n".join(pages)


def main():
    print("=" * 75)
    print("FULL ACCURACY CHECKER")
    print("Original 15 Scanned PDFs  vs  Excel/PDF File")
    print("=" * 75)
    print(f"Start time: {datetime.now().strftime('%d %B %Y, %I:%M %p')}\n")

    pdf_files = [f for f in os.listdir(SCANNED_PDF_FOLDER) if f.lower().endswith(".pdf")]
    pdf_files.sort()
    print(f"Found {len(pdf_files)} original PDF files\n")

    all_original_numbers = []

    print("STAGE 1: Processing the 15 original PDFs")
    print("-" * 75)

    for idx, pdf_name in enumerate(pdf_files, 1):
        pdf_path = os.path.join(SCANNED_PDF_FOLDER, pdf_name)
        txt_name = pdf_name.rsplit(".", 1)[0] + ".txt"
        txt_path = os.path.join(OCR_TEXT_FOLDER, txt_name)

        print(f"[{idx}/{len(pdf_files)}] {pdf_name}")

        if os.path.exists(txt_path):
            print("   → Using previously saved OCR text")
            with open(txt_path, "r", encoding="utf-8") as f:
                text = f.read()
        else:
            try:
                text = ocr_pdf(pdf_path, txt_path)
            except Exception as e:
                print(f"   → ERROR: {e}")
                text = ""

        numbers = extract_numbers(text)
        all_original_numbers.extend(numbers)

    unique_original_numbers = list(dict.fromkeys(all_original_numbers))
    print(f"\nTotal unique numbers found across 15 PDFs: {len(unique_original_numbers):,}")

    print("\nSTAGE 2: Processing the Excel-converted PDF")
    print("-" * 75)

    excel_text = extract_text_from_clean_pdf(EXCEL_PDF_PATH)
    excel_numbers = extract_numbers(excel_text)
    unique_excel_numbers = list(dict.fromkeys(excel_numbers))
    print(f"Unique numbers found in Excel PDF: {len(unique_excel_numbers):,}")

    original_set = set(unique_original_numbers)
    excel_set = set(unique_excel_numbers)

    matched = original_set.intersection(excel_set)
    missing_in_excel = original_set - excel_set
    extra_in_excel = excel_set - original_set

    matched_count = len(matched)
    missing_count = len(missing_in_excel)
    extra_count = len(extra_in_excel)

    if len(unique_original_numbers) > 0:
        accuracy = (matched_count / len(unique_original_numbers)) * 100
    else:
        accuracy = 0.0

    print("\n" + "=" * 75)
    print("FINAL RESULTS")
    print("=" * 75)
    print(f"Total unique numbers in 15 Original PDFs     : {len(unique_original_numbers):,}")
    print(f"Total unique numbers in Excel/PDF file       : {len(unique_excel_numbers):,}")
    print(f"Numbers successfully matched                 : {matched_count:,}")
    print(f"Numbers missing in Excel/PDF                 : {missing_count:,}")
    print(f"Numbers present only in Excel/PDF (extra)    : {extra_count:,}")
    print("-" * 75)
    print(f"MATCH RATE / ACCURACY                        : {accuracy:.2f}%")
    print("=" * 75)

    # Save reports
    report_path = os.path.join(OUTPUT_FOLDER, "Full_Accuracy_Report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("FULL ACCURACY REPORT\n")
        f.write("Original 15 Scanned PDFs vs Excel/PDF File\n")
        f.write("=" * 60 + "\n")
        f.write(f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}\n\n")
        f.write(f"Match Rate / Accuracy: {accuracy:.2f}%\n")
        f.write(f"Matched: {matched_count}\n")
        f.write(f"Missing: {missing_count}\n")
        f.write(f"Extra: {extra_count}\n")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Full Accuracy Report - 15 PDFs vs Excel/PDF"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Metric"
    ws["B3"] = "Count"
    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)

    ws["A4"] = "Unique numbers in 15 Original PDFs"
    ws["B4"] = len(unique_original_numbers)
    ws["A5"] = "Unique numbers in Excel/PDF"
    ws["B5"] = len(unique_excel_numbers)
    ws["A6"] = "Numbers Matched"
    ws["B6"] = matched_count
    ws["A7"] = "Numbers Missing in Excel/PDF"
    ws["B7"] = missing_count
    ws["A8"] = "Extra numbers only in Excel/PDF"
    ws["B8"] = extra_count

    ws["A10"] = "MATCH RATE / ACCURACY"
    ws["B10"] = f"{accuracy:.2f}%"
    ws["A10"].font = Font(bold=True, color="006600")
    ws["B10"].font = Font(bold=True, color="006600", size=14)

    ws2 = wb.create_sheet("Missing in Excel")
    ws2["A1"] = "Numbers present in Original 15 PDFs but missing in Excel/PDF"
    ws2["A1"].font = Font(bold=True)
    for i, num in enumerate(sorted(missing_in_excel, key=lambda x: float(x) if x.replace(".", "").isdigit() else 0), 2):
        ws2[f"A{i}"] = num

    ws3 = wb.create_sheet("Extra in Excel")
    ws3["A1"] = "Numbers present only in Excel/PDF"
    ws3["A1"].font = Font(bold=True)
    for i, num in enumerate(sorted(extra_in_excel, key=lambda x: float(x) if x.replace(".", "").isdigit() else 0), 2):
        ws3[f"A{i}"] = num

    excel_path = os.path.join(OUTPUT_FOLDER, "Full_Accuracy_Report.xlsx")
    wb.save(excel_path)

    print(f"\nReports saved in: {OUTPUT_FOLDER}")
    print(f"Final Match Rate / Accuracy: {accuracy:.2f}%")


if __name__ == "__main__":
    main()
