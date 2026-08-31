import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# ====================== SETTINGS ======================
ORIGINAL_FILE = r"C:\Users\padma\Documents\OCR_Comparison_Results\ALL_15_REPORTS_OCR_COMBINED.txt"
PROCESSED_FILE = r"C:\Users\padma\Documents\OCR_Comparison_Results\EXCEL_PDF_TEXT.txt"

OUTPUT_FOLDER = r"C:\Users\padma\Documents\Data_Difference_Report"
# ======================================================

os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def read_file(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        print(f"Error reading {path}: {e}")
        return ""


def get_basic_stats(text, label):
    char_count = len(text)
    char_count_no_space = len(text.replace(" ", "").replace("\n", "").replace("\t", ""))
    word_count = len(text.split())
    line_count = text.count("\n") + 1 if text else 0

    numbers = re.findall(r'\b\d{1,3}(?:,\d{3})*(?:\.\d+)?\b|\b\d+\.\d+\b|\b\d+\b', text)
    unique_numbers = list(dict.fromkeys(numbers))

    print(f"\n----- {label} -----")
    print(f"Total characters (with spaces) : {char_count:,}")
    print(f"Total characters (no spaces)   : {char_count_no_space:,}")
    print(f"Total words                    : {word_count:,}")
    print(f"Total lines                    : {line_count:,}")
    print(f"Unique numbers found           : {len(unique_numbers):,}")

    return {
        "char_count": char_count,
        "char_count_no_space": char_count_no_space,
        "word_count": word_count,
        "line_count": line_count,
        "unique_numbers": unique_numbers
    }


def find_missing_and_extra(original_numbers, processed_numbers):
    original_set = set(original_numbers)
    processed_set = set(processed_numbers)

    missing = list(original_set - processed_set)
    extra = list(processed_set - original_set)

    def sort_key(x):
        try:
            return float(x.replace(",", ""))
        except:
            return 0

    return sorted(missing, key=sort_key), sorted(extra, key=sort_key)


def main():
    print("=" * 70)
    print("DATA DIFFERENCE + ACCURACY CHECKER")
    print("Comparing Original Source vs Processed Excel/PDF file")
    print("=" * 70)
    print(f"Start time: {datetime.now().strftime('%d %B %Y, %I:%M %p')}\n")

    original_text = read_file(ORIGINAL_FILE)
    processed_text = read_file(PROCESSED_FILE)

    if not original_text or not processed_text:
        print("ERROR: One or both files could not be read.")
        return

    original_stats = get_basic_stats(original_text, "ORIGINAL FILE (Source)")
    processed_stats = get_basic_stats(processed_text, "PROCESSED FILE (Excel/PDF)")

    char_diff = original_stats["char_count"] - processed_stats["char_count"]
    word_diff = original_stats["word_count"] - processed_stats["word_count"]

    missing, extra = find_missing_and_extra(
        original_stats["unique_numbers"],
        processed_stats["unique_numbers"]
    )

    matched_count = len(original_stats["unique_numbers"]) - len(missing)

    if len(original_stats["unique_numbers"]) > 0:
        accuracy = (matched_count / len(original_stats["unique_numbers"])) * 100
    else:
        accuracy = 0.0

    print("\n" + "=" * 70)
    print("SUMMARY OF CHANGES + ACCURACY")
    print("=" * 70)
    print(f"Characters difference           : {char_diff:,}")
    print(f"Words difference                : {word_diff:,}")
    print(f"Unique numbers in Original      : {len(original_stats['unique_numbers']):,}")
    print(f"Unique numbers in Processed     : {len(processed_stats['unique_numbers']):,}")
    print(f"Numbers successfully matched    : {matched_count:,}")
    print(f"Numbers missing from Processed  : {len(missing):,}")
    print(f"Numbers newly added             : {len(extra):,}")
    print("-" * 70)
    print(f"OVERALL DATA ACCURACY           : {accuracy:.2f}%")
    print("=" * 70)

    # Text report
    report_path = os.path.join(OUTPUT_FOLDER, "Difference_and_Accuracy_Report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("DATA DIFFERENCE + ACCURACY REPORT\n")
        f.write("=" * 60 + "\n")
        f.write(f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}\n\n")
        f.write(f"Overall Accuracy: {accuracy:.2f}%\n")
        f.write(f"Matched numbers: {matched_count}\n")
        f.write(f"Missing numbers: {len(missing)}\n")
        f.write(f"Extra numbers: {len(extra)}\n")

    # Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Data Difference + Accuracy Report"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Metric"
    ws["B3"] = "Original"
    ws["C3"] = "Processed"
    ws["D3"] = "Difference"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}3"].font = Font(bold=True)

    ws["A4"] = "Characters"
    ws["B4"] = original_stats["char_count"]
    ws["C4"] = processed_stats["char_count"]
    ws["D4"] = char_diff

    ws["A5"] = "Words"
    ws["B5"] = original_stats["word_count"]
    ws["C5"] = processed_stats["word_count"]
    ws["D5"] = word_diff

    ws["A6"] = "Unique Numbers"
    ws["B6"] = len(original_stats["unique_numbers"])
    ws["C6"] = len(processed_stats["unique_numbers"])

    ws["A8"] = "Numbers Matched"
    ws["B8"] = matched_count
    ws["A9"] = "Numbers Missing"
    ws["B9"] = len(missing)
    ws["A10"] = "Numbers Extra"
    ws["B10"] = len(extra)

    ws["A12"] = "OVERALL DATA ACCURACY"
    ws["B12"] = f"{accuracy:.2f}%"
    ws["A12"].font = Font(bold=True, color="006600")
    ws["B12"].font = Font(bold=True, color="006600", size=14)

    ws2 = wb.create_sheet("Missing Numbers")
    ws2["A1"] = "Numbers missing in Processed file"
    ws2["A1"].font = Font(bold=True)
    for i, num in enumerate(missing, 2):
        ws2[f"A{i}"] = num

    ws3 = wb.create_sheet("Extra Numbers")
    ws3["A1"] = "Numbers only in Processed file"
    ws3["A1"].font = Font(bold=True)
    for i, num in enumerate(extra, 2):
        ws3[f"A{i}"] = num

    excel_path = os.path.join(OUTPUT_FOLDER, "Difference_and_Accuracy_Report.xlsx")
    wb.save(excel_path)

    print(f"\nReports saved in: {OUTPUT_FOLDER}")
    print(f"Overall Accuracy: {accuracy:.2f}%")


if __name__ == "__main__":
    main()
